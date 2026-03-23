from flask import Flask, request, render_template, redirect
import requests
import openpyxl
import os
import logging

app = Flask(__name__)

# ---------------- LOGGING ----------------
logging.basicConfig(level=logging.INFO)

# ---------------- API CONFIG ----------------
INSTANCE_KEY = "instance143653"
TOKEN = os.environ.get("TOKEN")   # FIXED
API_URL = f"https://api.ultramsg.com/{INSTANCE_KEY}/messages/chat"

# ---------------- EXCEL FILE ----------------
EXCEL_FILE = "Student_Data.xlsx"

# Create file if not exists
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Student Data"
    sheet.append(["Student Name", "Phone Number", "Course Name", "Parent Name", "Parent Contact"])
    workbook.save(EXCEL_FILE)

# ---------------- SAVE FUNCTION ----------------
def save_to_excel(data):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(EXCEL_FILE)

# ---------------- CHECK DUPLICATE ----------------
def is_duplicate(phone):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == phone:
            return True
    return False

# ---------------- WHATSAPP FUNCTION ----------------
def send_whatsapp_message(phone_number, message):
    if not TOKEN:
        return {"error": "Token not configured"}

    if not phone_number.startswith("+91"):
        phone_number = "+91" + phone_number.lstrip("0")

    payload = {"to": phone_number, "body": message}

    try:
        response = requests.post(f"{API_URL}?token={TOKEN}", json=payload, timeout=10)
        return response.json()
    except Exception as e:
        return {"error": str(e)}

# ---------------- HOME ----------------
@app.route("/")
def home():
    return render_template("home.html")

# ---------------- VISITOR FORM ----------------
@app.route("/visitor_form")
def visitor_form():
    return render_template("index.html")

# ---------------- SEND MESSAGE ----------------
@app.route("/send_message", methods=["POST"])
def send_message():
    student_name = request.form.get("student_name", "").strip()
    student_number = request.form.get("student_number", "").strip()
    course_name = request.form.get("course_name", "").strip()
    parent_name = request.form.get("parent_name", "").strip()
    parent_contact = request.form.get("parent_contact", "").strip()

    # Validation
    if not all([student_name, student_number, course_name, parent_name, parent_contact]):
        return "❌ All fields are required."

    if len(student_number) < 10:
        return "❌ Invalid phone number."

    if is_duplicate(student_number):
        return "⚠️ Visitor already exists."

    # Save
    save_to_excel([student_name, student_number, course_name, parent_name, parent_contact])
    logging.info(f"New visitor added: {student_name}")

    # Message
    message = f"""
Hello {student_name},

Welcome to Vikrant Group Of Institutions, Indore.

Thank you for visiting our campus.

Courses: Engineering, Management, Nursing, Pharmacy, Law

Scholarship:
https://www.vitm.edu.in/scholarship.html

Instagram:
https://www.instagram.com/vikrant.indore

Thanks
"""

    result = send_whatsapp_message(student_number, message)

    if "error" in result:
        return f"⚠️ Saved but message failed: {result['error']}"

    return redirect("/visitor_form?success=1")

# ---------------- BULK MESSAGE ----------------
@app.route("/bulk_message", methods=["GET", "POST"])
def bulk_message():
    if request.method == "POST":
        message = request.form.get("message")
        manual_input = request.form.get("manual_numbers", "")
        file = request.files.get("file")

        numbers = []

        # Excel input
        if file and file.filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            if "Phone Number" not in header:
                return "⚠️ 'Phone Number' column not found."

            idx = header.index("Phone Number")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[idx]:
                    numbers.append(str(row[idx]))

        # Manual input
        if manual_input.strip():
            for num in manual_input.replace(',', '\n').split('\n'):
                if num.strip():
                    numbers.append(num.strip())

        if not numbers:
            return "⚠️ No numbers found."

        failed = []

        for number in numbers:
            result = send_whatsapp_message(number, message)
            if "error" in result:
                failed.append(number)

        if failed:
            return f"❌ Failed for: {', '.join(failed)}"

        return "✅ Bulk messages sent successfully!"

    return render_template("bulk_message.html")

# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
def dashboard():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    total = sheet.max_row - 1
    return render_template("dashboard.html", total=total)

# ---------------- VIEW VISITORS ----------------
@app.route("/view_visitors")
def view_visitors():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    data = list(sheet.values)
    headers = data[0]
    rows = data[1:]

    return render_template("view.html", headers=headers, rows=rows)

# ---------------- ERROR ----------------
@app.errorhandler(404)
def not_found(e):
    return "<h1>Page Not Found</h1>", 404

# ---------------- RUN ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)  
